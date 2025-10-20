# koshtorys.py
import os
import re
import tkinter.filedialog as filedialog
import tkinter.messagebox as messagebox
import sys
import traceback

from openpyxl import load_workbook

# Импортируем необходимые функции из созданных модулей
try:
    from error_handler import log_error_koshtorys as log_error # Используем алиас
    from text_utils import number_to_ukrainian_text
except ImportError:
    print("Критическая ошибка: не удалось импортировать модули error_handler или text_utils для koshtorys.py")
    # Можно добавить заглушки или sys.exit()
    def log_error(exc_type, exc_value, exc_traceback, error_log="error.txt"): # Заглушка
        print(f"Ошибка (koshtorys log): {exc_value}")
        traceback.print_exception(exc_type, exc_value, exc_traceback)
        messagebox.showerror("Ошибка", f"Произошла ошибка (koshtorys): {str(exc_value)}")
    def number_to_ukrainian_text(amount): # Заглушка
        return f"{amount} грн (заглушка)"


# Константы для ячеек в кошторис.xlsx - теперь в виде шаблонов для нескольких договоров
KOSHTORYS_CELLS_TEMPLATE = {
    "назва_заходу": "D12",
    "адреса": "E14",
    "дата": "E15",
    "товар": "C{row}",
    "кількість": "H{row}",
    "ціна за одиницю": "J{row}",
    "разом": "K{row}",
    "всього_сума_словами": "B45",
    "сума_словами": "H6",
    "загальна_сума": "K41"  # Добавлена ячейка с общей суммой
}

# Определяем поля, которые должны быть числами в Excel
NUMERIC_FIELDS = ["кількість", "ціна за одиницю", "разом"]

# Начальная строка для первого договора
BASE_ROW = 32 # В вашем коде было 32, но в fill_koshtorys используется range(27,41) для суммирования. Уточните.
              # Если BASE_ROW = 32, то первый товар будет в C32, H32, J32, K32

# Добавляем константу для пути к файлу кошторис.xlsx
KOSHTORYS_PATH = "ШАБЛОН_кошторис.xlsx"


# Функция для преобразования строки в число для Excel
def convert_to_excel_number(value_str):
    if value_str is None: return 0.0 # Обработка None
    # Удаляем пробелы и заменяем запятые на точки
    clean_str = str(value_str).replace(" ", "").replace(",", ".")

    # Извлекаем только цифры и одну десятичную точку
    matches = re.search(r'(\d+\.?\d*)', clean_str)
    if matches:
        number_str = matches.group(1)
        try:
            # Преобразуем в число
            return float(number_str)
        except ValueError:
            pass # Если не удалось, вернем 0.0 или исходную строку, в зависимости от логики

    # Если не удалось преобразовать, возвращаем 0.0 (или можно вернуть value_str для отладки)
    return 0.0 # Важно для суммирования, чтобы не было ошибок типов


# Функция для получения текущих ячейки для конкретной строки
def get_cells_for_row(row):
    cells = {}
    for key, template in KOSHTORYS_CELLS_TEMPLATE.items():
        if "{row}" in template:
            cells[key] = template.format(row=row)
        else:
            cells[key] = template
    return cells

# Функция для заполнения файла кошторис.xlsx
def fill_koshtorys(document_blocks):
    try:
        # Проверяем, есть ли документы
        if not document_blocks:
            messagebox.showwarning("Увага", "Не додано жодного кошториса для заповнення.")
            return False

        # Проверяем существование файла KOSHTORYS_PATH
        current_koshtorys_path = get_koshtorys_path() # Используем getter
        if not os.path.exists(current_koshtorys_path):
            # Спрашиваем пользователя о местоположении файла
            koshtorys_file_selected = filedialog.askopenfilename(
                title=f"Оберіть файл шаблону кошторису (поточний: {current_koshtorys_path} не знайдено)",
                filetypes=[("Excel Files", "*.xlsx")]
            )
            if not koshtorys_file_selected:
                messagebox.showwarning("Увага", "Файл шаблону кошторису не обрано.")
                return False
            set_koshtorys_path(koshtorys_file_selected) # Обновляем путь через setter
            current_koshtorys_path = koshtorys_file_selected
        else:
            koshtorys_file = current_koshtorys_path


        # Загружаем существующий файл
        wb = load_workbook(current_koshtorys_path)
        ws = wb.active

        # Функция для определения, есть ли ПДВ в любом из договоров
        def check_pdv_status(blocks):
            for block_item in blocks: # block переименован в block_item во избежание конфликта с переменной из lambda
                entries = block_item.get("entries", {})
                pdv_entry = entries.get("пдв")
                if pdv_entry and hasattr(pdv_entry, 'get'):
                    pdv_text = pdv_entry.get().strip().lower()
                    if "пдв" in pdv_text and "без" not in pdv_text:
                        return "з ПДВ"
            return "без ПДВ"

        pdv_status = check_pdv_status(document_blocks)

        # Функция для безопасной записи в ячейку, учитывая возможные объединенные ячейки
        def safe_write_cell(cell_address, value, is_numeric=False):
            try:
                final_value = convert_to_excel_number(value) if is_numeric else value
                ws[cell_address] = final_value
            except AttributeError: # Обычно ошибка при записи в объединенную ячейку
                merged_found = False
                for merged_range in ws.merged_cells.ranges:
                    if cell_address in merged_range:
                        # Используем верхнюю левую ячейку объединенного диапазона
                        # print(f"Writing to merged cell {merged_range.start_cell.coordinate} for address {cell_address}")
                        ws[merged_range.start_cell.coordinate] = convert_to_excel_number(value) if is_numeric else value
                        merged_found = True
                        return
                if not merged_found:
                    log_error(None, f"Не вдалось записати значення в ячейку {cell_address} (можливо, не об'єднана належним чином або захищена)", None)
                    # messagebox.showwarning("Увага", f"Не вдалось записати значення в ячейку {cell_address}")

        # Функция для безопасного чтения из ячейки
        def safe_read_cell(cell_address):
            try:
                return ws[cell_address].value
            except Exception: # Более общее исключение для объединенных ячеек
                for merged_range in ws.merged_cells.ranges:
                    if cell_address in merged_range:
                        return ws[merged_range.start_cell.coordinate].value
                return None


        if document_blocks:
            first_block_entries = document_blocks[0].get("entries", {})
            # Получаем ячейки с общими данными (используя BASE_ROW, но эти ячейки не {row} зависимы)
            common_cells = get_cells_for_row(get_base_row()) # Используем getter

            # Извлекаем значения из виджетов Entry, если они есть
            def get_entry_value(entry_map, key, default=""):
                widget = entry_map.get(key)
                return widget.get() if hasattr(widget, 'get') else default

            safe_write_cell(KOSHTORYS_CELLS_TEMPLATE["назва_заходу"], get_entry_value(first_block_entries, "захід"))
            safe_write_cell(KOSHTORYS_CELLS_TEMPLATE["адреса"], get_entry_value(first_block_entries, "адреса"))
            safe_write_cell(KOSHTORYS_CELLS_TEMPLATE["дата"], get_entry_value(first_block_entries, "дата"))

            current_base_row = get_base_row() # Используем getter
            for i, block_item in enumerate(document_blocks):
                entries = block_item.get("entries", {})
                current_row_for_koshtorys = current_base_row + i
                cells_for_this_item = get_cells_for_row(current_row_for_koshtorys)

                # Заполняем данные для текущего договора
                safe_write_cell(cells_for_this_item["товар"], get_entry_value(entries, "товар"))

                for field in NUMERIC_FIELDS: # NUMERIC_FIELDS из koshtorys.py
                    if field in cells_for_this_item and field in entries:
                        safe_write_cell(cells_for_this_item[field], get_entry_value(entries, field), is_numeric=True)

            # Суммируем значения в ячейках от K (BASE_ROW) до K (BASE_ROW + кол-во договоров -1)
            # Но ваш код суммирует K27:K40. Нужно решить, какой диапазон правильный.
            # Предположим, что K41 - это итог, который должен быть рассчитан формулой в Excel или здесь.
            # Если мы заполняем строки товаров начиная с BASE_ROW, то и суммировать нужно их.

            # Пересчет суммы на основе заполненных данных
            total_sum_calculated = 0.0
            # Строки, которые мы заполнили товарами
            start_sum_row = current_base_row
            end_sum_row = current_base_row + len(document_blocks)

            for r_idx in range(start_sum_row, end_sum_row):
                cell_k_address = f"K{r_idx}" # Ячейка "разом" для текущего товара
                cell_value = safe_read_cell(cell_k_address) # Читаем то, что только что записали
                if cell_value is not None:
                    # convert_to_excel_number вернет float
                    total_sum_calculated += convert_to_excel_number(cell_value)

            # Записываем вычисленную общую сумму в ячейку KOSHTORYS_CELLS_TEMPLATE["загальна_сума"] (K49)
            safe_write_cell(KOSHTORYS_CELLS_TEMPLATE["загальна_сума"], total_sum_calculated, is_numeric=True)

            # ===  НАЧАЛО ИЗМЕНЕНИЙ ДЛЯ ЗАПИСИ В B45  ===
            # Читаем значение из K49 (ранее K41)
            total_sum_from_excel = safe_read_cell(KOSHTORYS_CELLS_TEMPLATE["загальна_сума"])
            if total_sum_from_excel is not None:
                # Преобразуем значение в число (если это строка)
                total_sum_numeric = convert_to_excel_number(total_sum_from_excel)

                if total_sum_numeric > 0:
                    suma_propysom = number_to_ukrainian_text(total_sum_numeric)
                    suma_propysom = suma_propysom[0].upper() + suma_propysom[1:]

                    vsogo_text_b45 = f"Всього: ({suma_propysom}, {pdv_status})"
                    safe_write_cell(KOSHTORYS_CELLS_TEMPLATE["всього_сума_словами"], vsogo_text_b45)

                    vsogo_text_h6 = f"({suma_propysom}, {pdv_status})"
                    safe_write_cell(KOSHTORYS_CELLS_TEMPLATE["сума_словами"], vsogo_text_h6)
                else:
                    # Если сумма 0 или отрицательная
                    log_text = "Загальна сума з K49 дорівнює 0.0 або менше. Перевірте дані."
                    print(log_text)
                    # Оставляем поля B45 и H6 пустыми или с заглушкой
                    safe_write_cell(KOSHTORYS_CELLS_TEMPLATE["всього_сума_словами"], f"Всього: (Нуль гривень 00 коп., {pdv_status})")
                    safe_write_cell(KOSHTORYS_CELLS_TEMPLATE["сума_словами"], f"(Нуль гривень 00 коп., {pdv_status})")
            else:
                log_text = "Не вдалося отримати загальну суму з K49."
                print(log_text)
                #  Оставляем поля B45 и H6 пустыми или с сообщением об ошибке
                safe_write_cell(KOSHTORYS_CELLS_TEMPLATE["всього_сума_словами"], "Помилка: Не вдалося розрахувати суму")
                safe_write_cell(KOSHTORYS_CELLS_TEMPLATE["сума_словами"], "Помилка")
            # ===  КОНЕЦ ИЗМЕНЕНИЙ ДЛЯ ЗАПИСИ В B45  ===

            # Сохраняем изменения
            # Определяем директорию исходного шаблона кошториса
            source_dir = os.path.dirname(current_koshtorys_path)
            output_filename = "Кошторис_заповнений.xlsx" # Новое имя, чтобы не перезаписывать шаблон
            save_path = os.path.join(source_dir, output_filename)


            wb.save(save_path)
            messagebox.showinfo("Успіх", f"Кошторис успішно заповнено та збережено як:\n{save_path}")
    except Exception as e:
        return


def set_koshtorys_path(path):
    """Установить новый путь к файлу кошториса"""
    global KOSHTORYS_PATH
    KOSHTORYS_PATH = path
    return KOSHTORYS_PATH

def get_koshtorys_path():
    """Получить текущий путь к файлу кошториса"""
    return KOSHTORYS_PATH

def set_base_row(row):
    """Установить базовую строку для начала заполнения договоров"""
    global BASE_ROW
    try:
        BASE_ROW = int(row)
    except ValueError:
        messagebox.showerror("Помилка", "Базовий рядок має бути числом.")
        return BASE_ROW # Возвращаем старое значение
    return BASE_ROW

def get_base_row():
    """Получить текущую базовую строку"""
    return BASE_ROW

def update_koshtorys_cells_template(cells_dict):
    """Обновить шаблоны соответствия ячеек в файле кошториса"""
    global KOSHTORYS_CELLS_TEMPLATE
    if isinstance(cells_dict, dict):
        KOSHTORYS_CELLS_TEMPLATE.update(cells_dict)
    else:
        messagebox.showerror("Помилка", "Шаблон комірок має бути словником.")
    return KOSHTORYS_CELLS_TEMPLATE

def update_numeric_fields(fields_list):
    """Обновить список полей, которые должны быть числами в Excel"""
    global NUMERIC_FIELDS
    if isinstance(fields_list, list):
        NUMERIC_FIELDS = fields_list
    else:
        messagebox.showerror("Помилка", "Список числових полів має бути списком.")
    return NUMERIC_FIELDS