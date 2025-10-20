import openpyxl
from pathlib import Path


def full_diagnostic_koshtorys(file_path):
    """ПОВНИЙ аналіз структури Кошторису - показує ВСІ колонки"""

    if not Path(file_path).exists():
        print(f"❌ Файл {file_path} не знайдено!")
        return

    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Беремо перший лист (незалежно від назви)
    ws = wb.worksheets[0]
    sheet_name = ws.title

    print("=" * 120)
    print(f"ПОВНИЙ АНАЛІЗ КОШТОРИСУ (лист: '{sheet_name}')")
    print("=" * 120)

    # 1. Назва заходу в D12
    print("\n📋 ОСНОВНА ІНФОРМАЦІЯ:")
    print(f"  D12 (Назва заходу): {ws['D12'].value}")

    # 2. ПОКАЗУЄМО ВСІ КОЛОНКИ з рядка 27 до "Разом"
    print(f"\n📦 ТОВАРИ (рядки 27+) - ВСІ КОЛОНКИ A-P:")
    print(
        f"{'Row':<5} {'A':<8} {'B':<15} {'C':<25} {'D':<10} {'E':<10} {'F':<10} {'G':<8} {'H':<10} {'I':<10} {'J':<10} {'K':<12} {'L':<12} {'M':<10} {'N':<10} {'O':<10} {'P':<10}")
    print("-" * 120)

    for row in range(27, min(50, ws.max_row + 1)):  # Показуємо до 50 рядка
        c_val = ws[f'C{row}'].value

        # Зупиняємось на "Разом"
        if c_val and isinstance(c_val, str) and 'Разом за кошторисом' in c_val:
            print(f"\n{'=' * 120}")
            print(f"Row {row}: РАЗОМ ЗА КОШТОРИСОМ (зупиняємось)")
            print(f"{'=' * 120}")

            # Показуємо цей рядок окремо
            row_data = []
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
                val = ws[f'{col}{row}'].value
                row_data.append(str(val)[:10] if val else '─')

            print(f"{row:<5} {' | '.join(row_data)}")

            # Показуємо наступні 3 рядки (там КЕКВ суми)
            print(f"\n📊 РЯДКИ З ПІДСУМКАМИ КЕКВ:")
            for next_row in range(row + 1, row + 4):
                row_data = []
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
                    val = ws[f'{col}{next_row}'].value
                    row_data.append(str(val)[:10] if val else '─')
                print(f"{next_row:<5} {' | '.join(row_data)}")

            break

        # Показуємо всі колонки для цього рядка
        row_data = []
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
            val = ws[f'{col}{row}'].value
            if val is None:
                row_data.append('─')
            else:
                # Обрізаємо довгі значення
                val_str = str(val)
                if len(val_str) > 12:
                    val_str = val_str[:9] + "..."
                row_data.append(val_str)

        # Виділяємо "Нагородна атрибутика"
        marker = " ⚠️ НАГОРОДНА" if (c_val and 'Нагородна' in str(c_val)) else ""

        print(
            f"{row:<5} {row_data[0]:<8} {row_data[1]:<15} {row_data[2]:<25} {row_data[3]:<10} {row_data[4]:<10} {row_data[5]:<10} {row_data[6]:<8} {row_data[7]:<10} {row_data[8]:<10} {row_data[9]:<10} {row_data[10]:<12} {row_data[11]:<12} {row_data[12]:<10} {row_data[13]:<10} {row_data[14]:<10} {row_data[15]:<10}{marker}")

    # 3. АНАЛІЗ: які колонки містять числові дані
    print(f"\n📊 АНАЛІЗ ЧИСЛОВИХ КОЛОНОК (рядки 27-45):")

    numeric_cols = {}
    for col_letter in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
        has_numbers = False
        sample_values = []

        for row in range(27, 46):
            val = ws[f'{col_letter}{row}'].value
            if val is not None and val != 0:
                if isinstance(val, (int, float)):
                    has_numbers = True
                    sample_values.append(f"{val}")
                elif isinstance(val, str):
                    try:
                        num = float(val.replace(',', '.'))
                        has_numbers = True
                        sample_values.append(f"{num}")
                    except:
                        pass

        if has_numbers:
            numeric_cols[col_letter] = sample_values[:3]  # Перші 3 зразки

    print("\nКолонки з числами:")
    for col, samples in numeric_cols.items():
        print(f"  Колонка {col}: {', '.join(samples)}")

    # 4. РЕКОМЕНДАЦІЯ
    print(f"\n" + "=" * 120)
    print("🎯 РЕКОМЕНДАЦІЇ:")
    print("=" * 120)
    print("""
    Подивись на таблицю вище і скажи:
    1. В якій колонці знаходяться СУМИ товарів? (K, L, або інша?)
    2. В якій колонці знаходиться КЕКВ? (G або інша?)
    3. В якій колонці знаходяться НАЗВИ товарів? (C або CDEF об'єднані?)

    Це допоможе мені виправити код read_koshtorys_data()!
    """)

    wb.close()


if __name__ == '__main__':
    try:
        full_diagnostic_koshtorys('Кошторис.xlsx')
    except Exception as e:
        print(f"\n❌ ПОМИЛКА: {e}")
        import traceback

        traceback.print_exc()
    finally:
        input("\nНатисніть Enter для закриття...")