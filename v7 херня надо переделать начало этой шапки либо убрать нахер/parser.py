import openpyxl
from pathlib import Path


def analyze_koshtorys(file_path):
    """Аналізує структуру Кошторису"""

    if not Path(file_path).exists():
        print(f"❌ Файл {file_path} не знайдено!")
        return

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    print("=" * 100)
    print("АНАЛІЗ СТРУКТУРИ КОШТОРИСУ")
    print("=" * 100)

    # 1. Назва заходу в D12
    print("\n📋 ОСНОВНА ІНФОРМАЦІЯ:")
    print(f"  D12 (Назва заходу): {ws['D12'].value}")

    # 2. Знаходимо початок товарів (рядок 27)
    print(f"\n📦 ТОВАРИ (починаючи з рядка 27):")
    print(f"{'Рядок':<8} {'C (назва)':<40} {'G (КЕКВ)':<10} {'K (сума)':<15} {'Примітка'}")
    print("-" * 100)

    tovar_rows = []
    razom_row = None
    nagrada_row = None

    for row in range(27, ws.max_row + 1):
        c_val = ws[f'C{row}'].value
        g_val = ws[f'G{row}'].value
        k_val = ws[f'K{row}'].value

        # Перевіряємо чи це "Разом за кошторисом"
        if c_val and isinstance(c_val, str) and 'Разом за кошторисом' in c_val:
            razom_row = row
            print(f"{row:<8} {str(c_val)[:40]:<40} {'─':<10} {str(k_val):<15} ⚠️  РАЗОМ - сума всіх")
            break

        # Перевіряємо чи це "Нагородна атрибутика"
        if c_val and isinstance(c_val, str) and 'Нагородна атрибутика' in c_val:
            nagrada_row = row
            print(f"{row:<8} {str(c_val)[:40]:<40} {str(g_val):<10} {str(k_val):<15} ⚠️  НЕ ТОВАР!")
            continue

        # Якщо є дані - це товар/послуга
        if c_val or g_val or k_val:
            tovar_rows.append({
                'row': row,
                'name': c_val,
                'kekv': g_val,
                'suma': k_val
            })

            note = ""
            if c_val and 'Нагородна' in str(c_val):
                note = "⚠️  НЕ ТОВАР!"

            print(f"{row:<8} {str(c_val)[:40]:<40} {str(g_val):<10} {str(k_val):<15} {note}")

    print(f"\n📊 ПІДСУМКИ ПО КЕКВ (розраховані з товарів):")
    print(f"{'КЕКВ':<10} {'Сума':<15} {'Деталі'}")
    print("-" * 100)

    kekv_data = {}

    # Групуємо товари по КЕКВ і рахуємо суми
    for tovar in tovar_rows:
        # Пропускаємо "Нагородна атрибутика"
        if tovar['name'] and 'Нагородна атрибутика' in str(tovar['name']):
            continue

        if tovar['kekv']:
            kekv_num = str(tovar['kekv']).strip()
            suma = tovar['suma'] or 0

            if isinstance(suma, str):
                try:
                    suma = float(suma)
                except:
                    suma = 0

            if kekv_num not in kekv_data:
                kekv_data[kekv_num] = 0

            kekv_data[kekv_num] += suma

    # Виводимо суми по КЕКВ
    for kekv in sorted(kekv_data.keys()):
        suma = kekv_data[kekv]
        print(f"{kekv:<10} {suma:<15.2f} КЕКВ {kekv}: {suma:.2f} грн")

    # Перевіряємо чи "Разом" співпадає з сумою товарів
    razom_suma = ws[f'K{razom_row}'].value if razom_row else 0
    total_tovarov = sum(kekv_data.values())

    if razom_row:
        print(f"\n  Перевірка: Разом за кошторисом (K{razom_row}): {razom_suma}")
        print(f"  Сума по КЕКВ: {total_tovarov:.2f}")
        print(f"  Співпадає: {'✓' if abs(razom_suma - total_tovarov) < 0.01 else '✗'}")

    print(f"\n✅ СТАТИСТИКА:")
    print(f"  Знайдено товарів: {len(tovar_rows)}")
    print(f"  Рядок 'Разом за кошторисом': {razom_row}")
    print(f"  Рядок 'Нагородна атрибутика': {nagrada_row}")
    print(f"  Перший товар: рядок {tovar_rows[0]['row'] if tovar_rows else 'N/A'}")
    print(f"  Останній товар: рядок {tovar_rows[-1]['row'] if tovar_rows else 'N/A'}")
    print(f"\n  КЕКВ ДАНІ:")
    for kekv, suma in kekv_data.items():
        print(f"    КЕКВ {kekv}: {suma}")

    print("\n" + "=" * 100)
    print("РЕКОМЕНДАЦІЇ ДЛЯ ОНОВЛЕННЯ read_koshtorys_data():")
    print("=" * 100)
    print("""
    1. Починати з рядка 27 (замість динамічного пошуку)
    2. Зупинятися коли знайдемо 'Разом за кошторисом'
    3. Пропускати рядки з 'Нагородна атрибутика'
    4. Читати суми по КЕКВ після рядка 'Разом'
    5. Дані розташовані в: C (назва), G (КЕКВ), K (сума)
    """)

    wb.close()

if __name__ == '__main__':
    try:
        analyze_koshtorys('Кошторис.xlsx')
    except Exception as e:
        print(f"\n❌ ПОМИЛКА: {e}")
        import traceback

        traceback.print_exc()
    finally:
        input("\nНатисніть Enter для закриття...")
# ЗАПУСК
