import openpyxl
from pathlib import Path
import re
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# КОНСТАНТИ для назв товарів-остачі
OSTACHA_NAMES = {
    '2210': 'Футболки з нанесенням логотипу',
    '2240': 'Перевезення',
    '2250': 'Добові'
}

# Колір для виділення остачі (блакитний)
OSTACHA_COLOR = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")


def calculate_formula_value(formula_str, row_num, ws):
    """Вичисляє значення формули типа =D2*F2 або =G3+G4"""
    if not formula_str or not isinstance(formula_str, str):
        return formula_str

    if not formula_str.startswith('='):
        return formula_str

    formula = formula_str[1:]

    # Підставляємо значення для поточного рядка
    for r in range(2, 15):
        for col_letter in ['D', 'E', 'F', 'G']:
            cell_val = ws[f'{col_letter}{r}'].value
            if isinstance(cell_val, str) and cell_val.startswith('='):
                try:
                    d_val = ws[f'D{r}'].value or 0
                    e_val = ws[f'E{r}'].value or 0
                    f_val = ws[f'F{r}'].value or 0
                    nested_formula = cell_val[1:]
                    nested_result = eval(nested_formula.replace(f'D{r}', str(d_val))
                                         .replace(f'E{r}', str(e_val))
                                         .replace(f'F{r}', str(f_val)))
                    formula = formula.replace(f'{col_letter}{r}', str(nested_result))
                except:
                    formula = formula.replace(f'{col_letter}{r}', '0')
            else:
                val = cell_val or 0
                formula = formula.replace(f'{col_letter}{r}', str(val))

    try:
        result = eval(formula)
        return result
    except:
        return 0


def get_cell_value(ws, cell_address):
    """Отримує вичислене значення комірки"""
    cell = ws[cell_address]
    if cell.value is not None and not isinstance(cell.value, str):
        return cell.value
    if isinstance(cell.value, str) and cell.value.startswith('='):
        if hasattr(cell, 'cached_value') and cell.cached_value is not None:
            return cell.cached_value
        return None
    return cell.value


def parse_formula_references(formula_str):
    """Витягує всі посилання на комірки з формули"""
    if not formula_str or not isinstance(formula_str, str):
        return []
    if formula_str.startswith('='):
        formula_str = formula_str[1:]
    pattern = r'[A-Z]+\d+'
    matches = re.findall(pattern, formula_str)
    return matches


def get_kekv_mapping_for_zahid(rozrahunok_file, zahid_start_row):
    """Визначає які рядки належать до яких КЕКВ на основі формул"""
    kekv_mapping = {}

    wb_formulas = openpyxl.load_workbook(rozrahunok_file, data_only=False)
    ws_formulas = wb_formulas.active

    kekv_columns = {'L': '2210', 'M': '2240', 'N': '2250'}

    for col_letter, kekv_code in kekv_columns.items():
        cell_address = f'{col_letter}{zahid_start_row}'
        cell = ws_formulas[cell_address]

        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            references = parse_formula_references(cell.value)
            for ref in references:
                row_match = re.search(r'\d+', ref)
                if row_match:
                    row_num = int(row_match.group())
                    if row_num not in kekv_mapping:
                        kekv_mapping[row_num] = []
                    kekv_mapping[row_num].append(kekv_code)

    wb_formulas.close()
    return kekv_mapping


def find_zahid_in_rozrahunok(ws_rozrahunok, zahid_number):
    """Знаходить захід в Розрахунку за номером"""
    for row in range(2, ws_rozrahunok.max_row + 1):
        a_val = ws_rozrahunok[f'A{row}'].value
        if a_val is not None and isinstance(a_val, (int, float)) and int(a_val) == zahid_number:
            b_val = ws_rozrahunok[f'B{row}'].value  # Назва
            c_val = ws_rozrahunok[f'C{row}'].value  # Термін

            # Визначаємо межі заходу
            end_row = row
            for next_row in range(row + 1, ws_rozrahunok.max_row + 1):
                next_a_val = ws_rozrahunok[f'A{next_row}'].value
                if next_a_val is not None and isinstance(next_a_val, (int, float)):
                    end_row = next_row - 1
                    break
            else:
                end_row = ws_rozrahunok.max_row

            return {
                'number': int(a_val),
                'name': b_val,
                'termin': c_val,
                'start_row': row,
                'end_row': end_row
            }
    return None


def get_zalushky_for_zahid(ws_rozrahunok, zahid_start_row):
    """Отримує залишки по КЕКВ для заходу з колонок T, U, V"""
    return {
        '2210': get_cell_value(ws_rozrahunok, f'T{zahid_start_row}') or 0,
        '2240': get_cell_value(ws_rozrahunok, f'U{zahid_start_row}') or 0,
        '2250': get_cell_value(ws_rozrahunok, f'V{zahid_start_row}') or 0
    }


def get_koshtorys_needs(koshtorys_data):
    """Розраховує потребу по кожному КЕКВ з Кошторису"""
    _, _, _, kekv_data = koshtorys_data

    needs = {
        '2210': kekv_data.get('2210', 0),
        '2240': kekv_data.get('2240', 0),
        '2250': kekv_data.get('2250', 0)
    }

    return needs


def find_closest_match(target, values, tolerance=100):
    """Знаходить найближче значення до цільового (шукає пару для підрахунку залишку)"""
    # Шукаємо значення, яке менше або дорівнює target і максимально близьке
    best_match = None
    min_diff = float('inf')

    for val in values:
        # Шукаємо використане значення, яке МЕНШЕ плану (або дорівнює)
        if val <= target:
            diff = abs(target - val)
            if diff <= tolerance and diff < min_diff:
                min_diff = diff
                best_match = val

    return best_match


def get_tovary_by_kekv(ws_rozrahunok, zahid, zalushky, kekv_mapping, allowed_kekv, rozrahunok_file):
    """Отримує список товарів з залишками (порівнює формули L/M/N з P/Q/R)"""
    tovary_by_kekv = {'2210': [], '2240': [], '2250': []}

    # Відкриваємо файл з формулами
    wb_formulas = openpyxl.load_workbook(rozrahunok_file, data_only=False)
    ws_formulas = wb_formulas.active

    # Колонки для кожного КЕКВ
    kekv_columns = {
        '2210': {'plan': 'L', 'used': 'P', 'zalushok': 'T'},
        '2240': {'plan': 'M', 'used': 'Q', 'zalushok': 'U'},
        '2250': {'plan': 'N', 'used': 'R', 'zalushok': 'V'}
    }

    for kekv in ['2210', '2240', '2250']:
        if kekv not in allowed_kekv:
            continue

        cols = kekv_columns[kekv]

        # Реальний залишок з T/U/V
        real_zalushok = zalushky.get(kekv, 0)
        if real_zalushok <= 0:
            continue

        print(f"\n  Аналіз КЕКВ {kekv}:")
        print(f"    Реальний залишок з {cols['zalushok']}: {real_zalushok}")

        # Збираємо ВСІ товари з формули плану L/M/N
        plan_cell = ws_formulas[f'{cols["plan"]}{zahid["start_row"]}']
        all_plan_items = {}  # {row: (name, suma)}

        if plan_cell.value and isinstance(plan_cell.value, str) and plan_cell.value.startswith('='):
            plan_refs = parse_formula_references(plan_cell.value)
            print(f"    План {cols['plan']}: {plan_cell.value}")

            for ref in plan_refs:
                row_match = re.search(r'\d+', ref)
                if row_match:
                    row_num = int(row_match.group())
                    h_val = ws_rozrahunok[f'H{row_num}'].value
                    k_val = get_cell_value(ws_rozrahunok, f'K{row_num}')
                    if h_val and k_val:
                        all_plan_items[row_num] = (h_val, k_val)
                        print(f"      - Рядок {row_num}: {h_val} = {k_val}")

        # Отримуємо значення "Використано" (P/Q/R)
        used_cell_formula = ws_formulas[f'{cols["used"]}{zahid["start_row"]}']
        used_cell_value = get_cell_value(ws_rozrahunok, f'{cols["used"]}{zahid["start_row"]}')

        used_sums = {}  # {suma_plan: suma_used} - пари для підрахунку залишків
        has_used_data = False

        # ВАРІАНТ 1: Є формула (наприклад =K5+K7)
        if used_cell_formula.value and isinstance(used_cell_formula.value, str) and used_cell_formula.value.startswith(
                '='):
            has_used_data = True
            formula = used_cell_formula.value[1:]  # Прибираємо =
            print(f"    Використано {cols['used']}: {used_cell_formula.value} (формула)")

            # Витягуємо всі використані суми
            used_values = set()
            refs = parse_formula_references(used_cell_formula.value)
            for ref in refs:
                row_match = re.search(r'\d+', ref)
                if row_match:
                    row_num = int(row_match.group())
                    k_val = get_cell_value(ws_rozrahunok, f'K{row_num}')
                    if k_val:
                        used_values.add(k_val)

            # Також шукаємо прості числа в формулі
            numbers = re.findall(r'\b(\d+\.?\d*)\b', formula)
            for num_str in numbers:
                try:
                    num = float(num_str)
                    used_values.add(num)
                except:
                    pass

            print(f"      Використані суми: {used_values}")

            # 🔥 КЛЮЧОВА ЛОГІКА: Шукаємо пари "план - використано"
            tolerance = 100  # Допуск для пошуку близьких значень (збільшено!)

            for row_num, (name, plan_suma) in all_plan_items.items():
                # Шукаємо найближче використане значення (яке МЕНШЕ плану)
                closest_used = find_closest_match(plan_suma, used_values, tolerance)

                if closest_used is not None:
                    # Знайдено пару: plan_suma ≈ closest_used
                    used_sums[plan_suma] = closest_used
                    diff = plan_suma - closest_used
                    print(f"      ✓ Пара: {plan_suma} (план) ≈ {closest_used} (використано), залишок = {diff:.2f}")
                else:
                    # Не знайдено використання - залишок = повна сума
                    used_sums[plan_suma] = 0
                    print(f"      ⚠️ Не знайдено використання для {plan_suma} (повний залишок)")

        # ВАРІАНТ 2: Є просте число (наприклад 1987.88)
        elif used_cell_value and isinstance(used_cell_value, (int, float)) and used_cell_value > 0:
            has_used_data = True
            print(f"    Використано {cols['used']}: {used_cell_value} (число)")

            # Розраховуємо яких товарів не вистачає до залишку
            plan_total = sum(suma for _, suma in all_plan_items.values())
            print(f"      План загалом: {plan_total:.2f}")
            print(f"      Використано: {used_cell_value:.2f}")
            print(f"      Залишок розрахунковий: {plan_total - used_cell_value:.2f}")

            # Шукаємо товари, які в сумі дають залишок
            tolerance = 0.1  # Допуск для порівняння

            # Перебираємо всі можливі комбінації товарів
            from itertools import combinations

            found_combination = False
            for r in range(1, len(all_plan_items) + 1):
                for combo in combinations(all_plan_items.items(), r):
                    combo_sum = sum(suma for _, (_, suma) in combo)
                    if abs(combo_sum - real_zalushok) < tolerance:
                        print(f"      ✓ Знайдено комбінацію товарів (залишок):")
                        for row_num, (name, suma) in combo:
                            print(f"        Рядок {row_num}: {name} = {suma:.2f}")
                            # ✅ КЛЮЧОВЕ: Ці товари ЗАЛИШИЛИСЬ (не використані повністю)
                            used_sums[suma] = 0  # Залишок = повна сума
                        found_combination = True
                        break
                if found_combination:
                    break

            # Якщо знайдено комбінацію залишку, решта товарів - використані
            if found_combination:
                for row_num, (name, suma) in all_plan_items.items():
                    if suma not in used_sums:
                        # Цей товар не в залишках = використано повністю
                        used_sums[suma] = suma  # Використано повністю
                        print(f"      ✗ Рядок {row_num}: {name} = {suma:.2f} (ВИКОРИСТАНО ПОВНІСТЮ)")

        # ВАРІАНТ 3: Порожньо або 0
        else:
            print(f"    Використано {cols['used']}: порожньо або 0")

        # 🎯 ВИЗНАЧАЄМО ЗАЛИШОК ТОВАРІВ
        print(f"    Товари що залишились:")

        if has_used_data:
            calculated_total = 0

            for row_num, (name, plan_suma) in all_plan_items.items():
                if plan_suma in used_sums:
                    used_suma = used_sums[plan_suma]
                    ostacha_tovaru = plan_suma - used_suma

                    if abs(ostacha_tovaru) < 0.01:  # Використано повністю (з округленням)
                        print(f"      ✗ Рядок {row_num}: {name} = {plan_suma:.2f} (ВИКОРИСТАНО ПОВНІСТЮ)")
                    elif ostacha_tovaru > 0.01:  # Є залишок
                        print(
                            f"      ✓ Рядок {row_num}: {name} = {plan_suma:.2f} - {used_suma:.2f} = {ostacha_tovaru:.2f}")
                        calculated_total += ostacha_tovaru

                        tovary_by_kekv[kekv].append({
                            'name': name,
                            'suma': ostacha_tovaru,
                            'kekv': kekv,
                            'row': row_num
                        })
                else:
                    # Не знайдено в used - значить залишок = повна сума
                    print(f"      ✓ Рядок {row_num}: {name} = {plan_suma:.2f} (НЕ ВИКОРИСТАНО)")
                    calculated_total += plan_suma

                    tovary_by_kekv[kekv].append({
                        'name': name,
                        'suma': plan_suma,
                        'kekv': kekv,
                        'row': row_num
                    })

            # Перевірка
            print(f"\n    ✓ ПЕРЕВІРКА: Розрахований залишок = {calculated_total:.2f}, Реальний = {real_zalushok:.2f}")

            if abs(calculated_total - real_zalushok) < 0.01:
                print(f"      ✅ ЗБІГАЄТЬСЯ!")
            else:
                diff = real_zalushok - calculated_total
                print(f"      ⚠️ РІЗНИЦЯ: {diff:.2f}")

                # Якщо різниця велика - запитуємо користувача
                if abs(diff) > 1:
                    print(f"\n      ⚠️ УВАГА: Розбіжність більше 1 грн!")
                    confirm = input(f"      Продовжити? (1-так, 0-скасувати): ").strip()
                    if confirm != '1':
                        print(f"      ✗ Скасовано користувачем")
                        tovary_by_kekv[kekv] = []
        else:
            # НЕМАЄ даних про використання - шукаємо товар з сумою = залишку
            print(f"    Шукаємо товар з сумою ≈ {real_zalushok}")
            tolerance = 0.01

            found = False
            for row_num, (name, suma) in all_plan_items.items():
                if abs(suma - real_zalushok) < tolerance:
                    print(f"      ✓ Рядок {row_num}: {name} = {suma} (СПІВПАДАЄ З ЗАЛИШКОМ)")
                    tovary_by_kekv[kekv].append({
                        'name': name,
                        'suma': suma,
                        'kekv': kekv,
                        'row': row_num
                    })
                    found = True
                else:
                    print(f"      ✗ Рядок {row_num}: {name} = {suma}")

            if not found:
                # Якщо не знайдено точного співпадіння
                print(f"      ⚠ Не знайдено товар з точною сумою залишку!")

                # Якщо є тільки ОДИН товар
                if len(all_plan_items) == 1:
                    row_num, (name, suma) = list(all_plan_items.items())[0]
                    print(f"\n      🤔 Знайдено ТІЛЬКИ ОДИН товар:")
                    print(f"         Рядок {row_num}: {name} = {suma:.2f}")
                    print(f"         Залишок: {real_zalushok:.2f}")
                    print(f"         Різниця: {abs(suma - real_zalushok):.2f}")

                    confirm = input(
                        f"\n      ❓ Залишок {real_zalushok:.2f} належить товару '{name}'?\n"
                        f"         1 - Так, підтверджую\n"
                        f"         2 - Ні, скасувати\n"
                        f"      👉 Ваш вибір: ").strip()

                    if confirm == '1':
                        print(f"      ✓ Підтверджено: переносимо '{name}'")
                        tovary_by_kekv[kekv].append({
                            'name': name,
                            'suma': real_zalushok,
                            'kekv': kekv,
                            'row': row_num
                        })
                    else:
                        print(f"      ✗ Скасовано користувачем")
                else:
                    # Кілька товарів
                    print(
                        f"      → Переносимо ВСІ товари (сума = {sum(suma for _, suma in all_plan_items.values()):.2f})")
                    for row_num, (name, suma) in all_plan_items.items():
                        print(f"      ✓ Рядок {row_num}: {name} = {suma} (ПЕРЕНОСИМО)")
                        tovary_by_kekv[kekv].append({
                            'name': name,
                            'suma': suma,
                            'kekv': kekv,
                            'row': row_num
                        })

    wb_formulas.close()
    return tovary_by_kekv


def read_koshtorys_data(ws_koshtorys):
    """Читає дані з Кошторису"""

    # Назва заходу в D12
    event_name = ws_koshtorys['D12'].value
    pp_number = "поточний захід з кошторису"

    print(f"\n📖 Читання Кошторису:")
    print(f"  п/п: {pp_number}")
    print(f"  Назва заходу: {event_name}")

    items = []
    razom_row = None

    # ЗНАХОДИМО ТОВАРИ (починаючи з рядка 27 до "Разом за кошторисом")
    print(f"\n📦 Обробка товарів:")

    for row in range(27, ws_koshtorys.max_row + 1):
        c_val = ws_koshtorys[f'C{row}'].value

        # Перевіряємо чи це "Разом за кошторисом"
        if c_val and isinstance(c_val, str) and 'Разом за кошторисом' in c_val:
            razom_row = row
            print(f"  ✓ Знайдено 'Разом за кошторисом' на рядку {row}")
            break

        # Перевіряємо чи це "Нагородна атрибутика" (це НЕ товар!)
        if c_val and isinstance(c_val, str) and 'Нагородна атрибутика' in c_val:
            print(f"  ⚠️  Рядок {row}: 'Нагородна атрибутика' (пропускаємо)")
            continue

        # Читаємо дані товару
        g_val = ws_koshtorys[f'G{row}'].value  # КЕКВ
        k_val = ws_koshtorys[f'K{row}'].value  # СУМА (основна колонка)

        # Якщо немає назви, КЕКВ або суми - пропускаємо
        if not c_val or not g_val or not k_val:
            continue

        # Конвертуємо в числа
        kekv = str(g_val).strip()

        suma = k_val
        if isinstance(suma, str):
            try:
                suma = float(suma.replace(',', '.').replace(' ', ''))
            except:
                suma = 0
        else:
            suma = float(suma) if suma else 0

        # Пропускаємо нульові суми
        if suma == 0:
            continue

        items.append({
            'name': str(c_val).strip(),
            'kekv': kekv,
            'suma': suma,
            'row': row
        })

        print(f"  [{row}] КЕКВ {kekv}: {c_val} = {suma:.2f}")

    # РОЗРАХОВУЄМО СУМИ ПО КЕКВ
    print(f"\n💰 Розрахунок сум по КЕКВ:")

    kekv_data = {}
    for item in items:
        kekv = item['kekv']
        if kekv not in kekv_data:
            kekv_data[kekv] = 0
        kekv_data[kekv] += item['suma']

    for kekv in sorted(kekv_data.keys()):
        suma = kekv_data[kekv]
        print(f"  КЕКВ {kekv}: {suma:.2f} грн")

    # ПЕРЕВІРКА з "Разом за кошторисом"
    if razom_row:
        razom_suma = ws_koshtorys[f'K{razom_row}'].value or 0
        if isinstance(razom_suma, str):
            try:
                razom_suma = float(razom_suma.replace(',', '.').replace(' ', ''))
            except:
                razom_suma = 0

        total_tovarov = sum(kekv_data.values())

        print(f"\n✓ Перевірка:")
        print(f"  'Разом за кошторисом' (K{razom_row}): {razom_suma:.2f}")
        print(f"  Сума по КЕКВ: {total_tovarov:.2f}")

        diff = abs(razom_suma - total_tovarov)
        if diff < 0.01:
            print(f"  Співпадає: ✓")
        else:
            print(f"  ⚠️  Різниця: {diff:.2f} грн")

    return pp_number, event_name, items, kekv_data


def find_zahid_by_name_in_rozrahunok(ws_rozrahunok, event_name):
    """Знаходить захід в Розрахунку за назвою"""
    if not event_name:
        return None

    # Очищуємо назву від лапок, пробілів та переносів рядків
    event_name_normalized = str(event_name).strip().lower()
    event_name_normalized = event_name_normalized.strip('"').strip("'").strip('«').strip('»').strip()
    event_name_normalized = event_name_normalized.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
    event_name_normalized = ' '.join(event_name_normalized.split())  # Прибираємо зайві пробіли

    # Спочатку шукаємо точне співпадіння


    for row in range(2, ws_rozrahunok.max_row + 1):
        b_val = ws_rozrahunok[f'B{row}'].value
        if b_val:
            b_val_normalized = str(b_val).strip().lower()
            b_val_normalized = b_val_normalized.strip('"').strip("'").strip('«').strip('»').strip()
            b_val_normalized = b_val_normalized.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
            b_val_normalized = ' '.join(b_val_normalized.split())

            if event_name_normalized == b_val_normalized:

                a_val = ws_rozrahunok[f'A{row}'].value
                c_val = ws_rozrahunok[f'C{row}'].value

                if a_val is not None and isinstance(a_val, (int, float)):
                    return {
                        'number': int(a_val),
                        'termin': c_val
                    }

    # Якщо точного співпадіння немає, шукаємо за спільними словами


    event_words = set(word for word in event_name_normalized.split() if len(word) >= 3)
    best_match = None
    max_common_words = 0

    for row in range(2, ws_rozrahunok.max_row + 1):
        b_val = ws_rozrahunok[f'B{row}'].value
        if b_val:
            b_val_normalized = str(b_val).strip().lower()
            b_val_normalized = b_val_normalized.strip('"').strip("'").strip('«').strip('»').strip()
            b_val_normalized = b_val_normalized.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
            b_val_normalized = ' '.join(b_val_normalized.split())

            b_words = set(word for word in b_val_normalized.split() if len(word) >= 3)
            common_words = event_words & b_words
            num_common = len(common_words)

            if num_common > max_common_words and num_common >= 3:
                max_common_words = num_common
                best_match = {
                    'row': row,
                    'name': b_val_normalized,
                    'common_words': common_words
                }

    if best_match:
        row = best_match['row']

        a_val = ws_rozrahunok[f'A{row}'].value
        c_val = ws_rozrahunok[f'C{row}'].value

        if a_val is not None and isinstance(a_val, (int, float)):
            return {
                'number': int(a_val),
                'termin': c_val
            }

    return None


def insert_data_to_zminy(zminy_file, koshtorys_data, zahody_rozrahunok, ostacha_items, rozrahunok_file):
    """Вставляє дані в Зміни.xlsx"""

    _, event_name, items_koshtorys, kekv_data = koshtorys_data

    # ✅ СТИЛІ: Границі + Шрифт
    border_style = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    times_font = Font(name='Times New Roman', size=11)
    times_font_bold = Font(name='Times New Roman', size=11, bold=True)

    # Відкриваємо або створюємо файл
    if zminy_file.exists():
        wb_zminy = openpyxl.load_workbook(zminy_file)
        ws_zminy = wb_zminy.active
    else:
        wb_zminy = openpyxl.Workbook()
        ws_zminy = wb_zminy.active
        # Створюємо заголовки
        headers = ['п/п', 'Назва заходу', 'Термін, місце проведення', 'Найменування',
                   'Сума', 'КЕКВ 2210', 'КЕКВ 2240', 'КЕКВ 2250', 'Сума витрат']

        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col, header in enumerate(headers, start=1):
            cell = ws_zminy.cell(row=1, column=col, value=header)
            cell.font = times_font  # ✅ Змінено: без жирного
            cell.alignment = center_alignment
            cell.border = border_style

        # Ширини колонок
        ws_zminy.column_dimensions['A'].width = 10
        ws_zminy.column_dimensions['B'].width = 20
        ws_zminy.column_dimensions['C'].width = 20
        ws_zminy.column_dimensions['D'].width = 23
        ws_zminy.column_dimensions['E'].width = 12
        ws_zminy.column_dimensions['F'].width = 15
        ws_zminy.column_dimensions['G'].width = 15
        ws_zminy.column_dimensions['H'].width = 15
        ws_zminy.column_dimensions['I'].width = 15

    # Зберігаємо старі дані (з 2-го рядка)
    old_data = []
    for row in range(2, ws_zminy.max_row + 1):
        row_data = []
        has_data = False
        for col in range(1, 10):
            val = ws_zminy.cell(row=row, column=col).value
            row_data.append(val)
            if val is not None:
                has_data = True
        if has_data:
            old_data.append(row_data)

    # Очищаємо все крім заголовків
    for row in ws_zminy.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill()
            cell.border = Border()

    current_row = 2
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # ✅ ФОРМАТ ЧИСЛА: пробіл як роздільник тисяч + 2 десяткові
    number_format = '#,##0.00'

    # ========== СПОЧАТКУ ЗАХОДИ З РОЗРАХУНКУ (З МІНУСОМ!) ==========

    for zahid_data in zahody_rozrahunok:
        zahid = zahid_data['zahid']
        tovary = zahid_data['tovary']
        zalushky = zahid_data.get('zalushky', {})

        # Рахуємо суми по КЕКВ для цього заходу (З МІНУСОМ!)
        kekv_sums = {
            '2210': -zalushky.get('2210', 0) if tovary['2210'] else 0,
            '2240': -zalushky.get('2240', 0) if tovary['2240'] else 0,
            '2250': -zalushky.get('2250', 0) if tovary['2250'] else 0
        }
        total_zahid = sum(kekv_sums.values())

        # Перевіряємо чи є товари взагалі
        if total_zahid == 0:
            continue

        # Запам'ятовуємо початок заходу
        zahid_start_row = current_row

        # Основний рядок заходу з Розрахунку
        ws_zminy.cell(row=current_row, column=1, value=zahid['number'])
        ws_zminy.cell(row=current_row, column=2, value=zahid['name'])
        ws_zminy.cell(row=current_row, column=3, value=zahid['termin'])

        # Суми КЕКВ (З МІНУСОМ!)
        ws_zminy.cell(row=current_row, column=6, value=kekv_sums['2210'])
        ws_zminy.cell(row=current_row, column=7, value=kekv_sums['2240'])
        ws_zminy.cell(row=current_row, column=8, value=kekv_sums['2250'])
        ws_zminy.cell(row=current_row, column=9, value=total_zahid)

        # Форматування
        for col in range(1, 10):
            cell = ws_zminy.cell(row=current_row, column=col)
            cell.alignment = center_alignment
            cell.border = border_style
            cell.font = times_font
            if col >= 6:
                cell.number_format = number_format

        # Перший товар в той же рядок
        first_added = False
        for kekv in ['2210', '2240', '2250']:
            if tovary[kekv] and not first_added:
                first_tovar = tovary[kekv][0]
                ws_zminy.cell(row=current_row, column=4, value=first_tovar['name'])
                ws_zminy.cell(row=current_row, column=5, value=-first_tovar['suma'])

                ws_zminy.cell(row=current_row, column=4).font = times_font
                ws_zminy.cell(row=current_row, column=5).font = times_font
                ws_zminy.cell(row=current_row, column=5).number_format = number_format

                # Видаляємо перший товар зі списку
                tovary[kekv].pop(0)
                first_added = True
                break

        current_row += 1

        # Решта товарів цього заходу
        for kekv in ['2210', '2240', '2250']:
            if tovary[kekv]:
                for tovar in tovary[kekv]:
                    ws_zminy.cell(row=current_row, column=4, value=tovar['name'])
                    ws_zminy.cell(row=current_row, column=5, value=-tovar['suma'])

                    for col in range(1, 10):
                        cell = ws_zminy.cell(row=current_row, column=col)
                        cell.alignment = center_alignment
                        cell.border = border_style
                        cell.font = times_font

                    ws_zminy.cell(row=current_row, column=5).number_format = number_format
                    current_row += 1

        # Запам'ятовуємо кінець заходу
        zahid_end_row = current_row - 1

        # ✅ ОБ'ЄДНАННЯ КОМІРОК ДЛЯ ЗАХОДУ З РОЗРАХУНКУ
        if zahid_end_row > zahid_start_row:
            merge_columns = [1, 2, 3, 6, 7, 8, 9]

            for col in merge_columns:
                ws_zminy.merge_cells(
                    start_row=zahid_start_row,
                    start_column=col,
                    end_row=zahid_end_row,
                    end_column=col
                )
                # Застосовуємо границі до всіх комірок у об'єднаному діапазоні
                for row in range(zahid_start_row, zahid_end_row + 1):
                    ws_zminy.cell(row=row, column=col).border = border_style

    # ========== ТЕПЕР ДОДАЄМО ДАНІ З КОШТОРИСУ (БЕЗ МІНУСА) ==========

    # Запам'ятовуємо початок заходу з Кошторису
    koshtorys_start_row = current_row

    # ✅ ШУКАЄМО ЗАХІД В РОЗРАХУНКУ ЗА НАЗВОЮ
    wb_rozrahunok_for_search = openpyxl.load_workbook(rozrahunok_file, data_only=True)
    ws_rozrahunok_for_search = wb_rozrahunok_for_search.active

    zahid_info = find_zahid_by_name_in_rozrahunok(ws_rozrahunok_for_search, event_name)
    wb_rozrahunok_for_search.close()

    # Ініціалізуємо змінні значеннями за замовчуванням
    pp_number = None
    termin = None

    if zahid_info:
        pp_number = zahid_info['number']
        termin = zahid_info['termin']
    else:
        print(f"\n⚠️  Захід '{event_name}' не знайдено в Розрахунку")
        print(f"   Стовпці A та C залишаться порожніми")

    # Рядок з основною інформацією заходу з Кошторису
    ws_zminy.cell(row=current_row, column=1, value=pp_number)
    ws_zminy.cell(row=current_row, column=2, value=event_name)
    ws_zminy.cell(row=current_row, column=3, value=termin)

    # Суми КЕКВ з Кошторису + ОСТАЧА!
    total_sum = sum(kekv_data.values()) + sum(ostacha_items.values())
    ws_zminy.cell(row=current_row, column=6, value=kekv_data.get('2210', 0) + ostacha_items.get('2210', 0))
    ws_zminy.cell(row=current_row, column=7, value=kekv_data.get('2240', 0) + ostacha_items.get('2240', 0))
    ws_zminy.cell(row=current_row, column=8, value=kekv_data.get('2250', 0) + ostacha_items.get('2250', 0))
    ws_zminy.cell(row=current_row, column=9, value=total_sum)

    # Форматування
    for col in range(1, 10):
        cell = ws_zminy.cell(row=current_row, column=col)
        cell.alignment = center_alignment
        cell.border = border_style
        cell.font = times_font
        if col >= 6:
            cell.number_format = number_format

    # Перший товар з Кошторису в той же рядок
    if items_koshtorys:
        first_item = items_koshtorys[0]
        ws_zminy.cell(row=current_row, column=4, value=first_item['name'])
        ws_zminy.cell(row=current_row, column=5, value=first_item['suma'])

        ws_zminy.cell(row=current_row, column=4).font = times_font
        ws_zminy.cell(row=current_row, column=5).font = times_font
        ws_zminy.cell(row=current_row, column=5).number_format = number_format

        current_row += 1

        # Решта товарів з Кошторису
        for item in items_koshtorys[1:]:
            ws_zminy.cell(row=current_row, column=4, value=item['name'])
            ws_zminy.cell(row=current_row, column=5, value=item['suma'])

            for col in range(1, 10):
                cell = ws_zminy.cell(row=current_row, column=col)
                cell.alignment = center_alignment
                cell.border = border_style
                cell.font = times_font

            ws_zminy.cell(row=current_row, column=5).number_format = number_format
            current_row += 1

    # Додаємо рядки остачі (якщо є) - ВИДІЛЯЄМО БЛАКИТНИМ
    for kekv, ostacha_suma in ostacha_items.items():
        if ostacha_suma > 0:
            ws_zminy.cell(row=current_row, column=4, value=OSTACHA_NAMES[kekv])
            ws_zminy.cell(row=current_row, column=5, value=ostacha_suma)

            # Форматування всіх комірок рядка
            for col in range(1, 10):
                cell = ws_zminy.cell(row=current_row, column=col)
                cell.alignment = center_alignment
                cell.border = border_style
                cell.font = times_font

            # Виділяємо блакитним
            ws_zminy.cell(row=current_row, column=4).fill = OSTACHA_COLOR
            ws_zminy.cell(row=current_row, column=5).fill = OSTACHA_COLOR
            ws_zminy.cell(row=current_row, column=5).number_format = number_format

            current_row += 1

    # Запам'ятовуємо кінець заходу з Кошторису
    koshtorys_end_row = current_row - 1

    # ✅ ОБ'ЄДНАННЯ КОМІРОК ДЛЯ ЗАХОДУ З КОШТОРИСУ
    if koshtorys_end_row > koshtorys_start_row:
        merge_columns = [1, 2, 3, 6, 7, 8, 9]

        for col in merge_columns:
            ws_zminy.merge_cells(
                start_row=koshtorys_start_row,
                start_column=col,
                end_row=koshtorys_end_row,
                end_column=col
            )
            # Застосовуємо границі до всіх комірок у об'єднаному діапазоні
            for row in range(koshtorys_start_row, koshtorys_end_row + 1):
                ws_zminy.cell(row=row, column=col).border = border_style

    # Повертаємо старі дані
    for old_row_data in old_data:
        for col, value in enumerate(old_row_data, start=1):
            cell = ws_zminy.cell(row=current_row, column=col, value=value)
            cell.border = border_style
            cell.font = times_font
            cell.alignment = center_alignment
            # Застосовуємо форматування до числових стовпців
            if col in [5, 6, 7, 8, 9] and isinstance(value, (int, float)):
                cell.number_format = number_format
        current_row += 1

    # Зберігаємо файл
    wb_zminy.save(zminy_file)
    wb_zminy.close()

    print(f"\n✅ Дані успішно записано в {zminy_file}")
    print(f"   Всього рядків: {current_row - 1}")



def choose_kekv_for_zahid(zahid, zalushky, koshtorys_needs):
    """Діалог вибору КЕКВ для конкретного заходу"""

    print("\n" + "=" * 80)
    print(f"📋 ЗАХІД #{zahid['number']}: {zahid['name']}")
    print("=" * 80)

    # Показуємо залишки по КЕКВ
    print(f"\n💰 ЗАЛИШКИ на цьому заході:")
    total_zalushky = 0
    for kekv in ['2210', '2240', '2250']:
        suma = zalushky.get(kekv, 0)
        if suma > 0:
            print(f"   КЕКВ {kekv}: {suma:>10.2f} грн")
            total_zalushky += suma

    if total_zalushky == 0:
        print(f"   ⚠️  Немає залишків на цьому заході!")
        return []

    print(f"   {'─' * 30}")
    print(f"   РАЗОМ:      {total_zalushky:>10.2f} грн")

    # Показуємо потреби з Кошторису
    print(f"\n📊 ПОТРЕБИ згідно Кошторису:")
    total_needs = 0
    for kekv in ['2210', '2240', '2250']:
        suma = koshtorys_needs.get(kekv, 0)
        if suma > 0:
            print(f"   КЕКВ {kekv}: {suma:>10.2f} грн")
            total_needs += suma

    if total_needs > 0:
        print(f"   {'─' * 30}")
        print(f"   РАЗОМ:      {total_needs:>10.2f} грн")

    # Порівняння
    if total_zalushky > 0 and total_needs > 0:
        diff = total_zalushky - total_needs
        print(f"\n⚖️  БАЛАНС:")
        if diff >= 0:
            print(f"   Залишків достатньо! Лишиться: {diff:.2f} грн")
        else:
            print(f"   ⚠️  Не вистачає: {abs(diff):.2f} грн")

    # Вибір КЕКВ
    print(f"\n🎯 З яких КЕКВ забирати залишки?")
    print(f"   Доступні опції:")

    available_kekv = []
    option_map = {}
    option_num = 1

    # Окремі КЕКВ
    for kekv in ['2210', '2240', '2250']:
        if zalushky.get(kekv, 0) > 0:
            print(f"   {option_num} - Тільки КЕКВ {kekv} ({zalushky[kekv]:.2f} грн)")
            option_map[str(option_num)] = [kekv]
            available_kekv.append(kekv)
            option_num += 1

    # Комбінації з 2-х КЕКВ
    if len(available_kekv) >= 2:
        for i in range(len(available_kekv)):
            for j in range(i + 1, len(available_kekv)):
                kekv1, kekv2 = available_kekv[i], available_kekv[j]
                suma = zalushky[kekv1] + zalushky[kekv2]
                print(f"   {option_num} - КЕКВ {kekv1} + {kekv2} ({suma:.2f} грн)")
                option_map[str(option_num)] = [kekv1, kekv2]
                option_num += 1

    # Всі КЕКВ
    if len(available_kekv) >= 2:
        print(f"   {option_num} - Всі КЕКВ ({total_zalushky:.2f} грн)")
        option_map[str(option_num)] = available_kekv
        option_num += 1

    # Пропустити
    print(f"   0 - Пропустити цей захід (не забирати залишки)")

    while True:
        choice = input(f"\n👉 Ваш вибір (0-{option_num - 1}): ").strip()

        if choice == '0':
            print(f"   ⏭️  Захід #{zahid['number']} пропущено")
            return []

        if choice in option_map:
            selected_kekv = option_map[choice]
            selected_suma = sum(zalushky[k] for k in selected_kekv)

            print(f"\n   ✓ Обрано: {', '.join(selected_kekv)}")
            print(f"   💵 Сума: {selected_suma:.2f} грн")

            return selected_kekv
        else:
            print(f"   ❌ Некоректний вибір! Спробуйте ще раз.")


def main():
    """Головна функція програми"""

    rozrahunok_file = Path('Розрахунок.xlsx')
    koshtorys_file = Path('Кошторис.xlsx')
    zminy_file = Path('Зміни.xlsx')

    print("=" * 80)
    print("🚀 ПРОГРАМА ПЕРЕНОСУ ДАНИХ З РОЗРАХУНКУ ТА КОШТОРИСУ В ЗМІНИ")
    print("=" * 80)

    # Перевірка наявності файлів
    if not rozrahunok_file.exists():
        print(f"\n❌ ПОМИЛКА: Файл '{rozrahunok_file}' не знайдено!")
        return

    if not koshtorys_file.exists():
        print(f"\n❌ ПОМИЛКА: Файл '{koshtorys_file}' не знайдено!")
        return

    # КРОК 1: Запитуємо номери заходів з Розрахунку
    print("\n" + "=" * 80)
    print("КРОК 1: Введення номерів заходів з Розрахунку")
    print("=" * 80)

    zahid_numbers_input = input("\nВведіть номери заходів через кому (наприклад: 1,3,5,9): ").strip()

    if not zahid_numbers_input:
        print("\n⚠️  Не введено жодного номера заходу!")
        return

    try:
        zahid_numbers = [int(x.strip()) for x in zahid_numbers_input.split(',')]
        print(f"✓ Будемо обробляти заходи: {zahid_numbers}")
    except ValueError:
        print("\n❌ ПОМИЛКА: Введено некоректні номери заходів!")
        return

    # КРОК 2: Читаємо Кошторис (щоб знати потреби)
    print("\n" + "=" * 80)
    print("КРОК 2: Аналіз Кошторису (визначення потреб)")
    print("=" * 80)

    wb_koshtorys = openpyxl.load_workbook(koshtorys_file, data_only=True)
    ws_koshtorys = wb_koshtorys.active

    koshtorys_data = read_koshtorys_data(ws_koshtorys)
    koshtorys_needs = get_koshtorys_needs(koshtorys_data)

    wb_koshtorys.close()

    print(f"\n📊 Загальні потреби з Кошторису:")
    total_need = 0
    for kekv in ['2210', '2240', '2250']:
        if koshtorys_needs[kekv] > 0:
            print(f"   КЕКВ {kekv}: {koshtorys_needs[kekv]:.2f} грн")
            total_need += koshtorys_needs[kekv]
    print(f"   РАЗОМ: {total_need:.2f} грн")

    # КРОК 3: Обробка кожного заходу окремо
    print("\n" + "=" * 80)
    print("КРОК 3: Вибір КЕКВ для кожного заходу")
    print("=" * 80)

    wb_rozrahunok = openpyxl.load_workbook(rozrahunok_file, data_only=True)
    ws_rozrahunok = wb_rozrahunok.active

    zahody_rozrahunok = []
    total_ostacha = {'2210': 0, '2240': 0, '2250': 0}

    for zahid_num in zahid_numbers:
        # Знаходимо захід
        zahid = find_zahid_in_rozrahunok(ws_rozrahunok, zahid_num)

        if not zahid:
            print(f"\n⚠️  Захід #{zahid_num} не знайдено в Розрахунку!")
            continue

        # Отримуємо залишки
        zalushky = get_zalushky_for_zahid(ws_rozrahunok, zahid['start_row'])

        # Діалог вибору КЕКВ для цього заходу
        selected_kekv = choose_kekv_for_zahid(zahid, zalushky, koshtorys_needs)

        if not selected_kekv:
            continue  # Користувач пропустив цей захід

        # Отримуємо товари тільки для обраних КЕКВ
        print(f"\n🔄 Обробка товарів заходу #{zahid_num}...")

        tovary = get_tovary_by_kekv(
            ws_rozrahunok,
            zahid,
            zalushky,
            {},
            selected_kekv,  # Передаємо тільки обрані КЕКВ!
            rozrahunok_file
        )

        # Додаємо до загальної остачі
        for kekv in selected_kekv:
            total_ostacha[kekv] += zalushky[kekv]

        zahody_rozrahunok.append({
            'zahid': zahid,
            'tovary': tovary,
            'zalushky': {k: zalushky[k] for k in selected_kekv}  # Зберігаємо залишки!
        })

    wb_rozrahunok.close()

    # КРОК 4: Підсумки
    print("\n" + "=" * 80)
    print("КРОК 4: ПІДСУМОК")
    print("=" * 80)

    print(f"\n✅ Оброблено заходів: {len(zahody_rozrahunok)}")

    if len(zahody_rozrahunok) == 0:
        print(f"\n⚠️  Не було обрано жодного заходу для переносу!")
        return

    print(f"\n💰 ЗАГАЛЬНА ОСТАЧА по всіх заходах:")
    total_ostacha_suma = 0
    for kekv in ['2210', '2240', '2250']:
        if total_ostacha[kekv] > 0:
            print(f"   КЕКВ {kekv}: {total_ostacha[kekv]:.2f} грн")
            total_ostacha_suma += total_ostacha[kekv]

    print(f"   {'─' * 30}")
    print(f"   РАЗОМ:      {total_ostacha_suma:.2f} грн")

    print(f"\n📊 ПОТРЕБИ з Кошторису:")
    for kekv in ['2210', '2240', '2250']:
        if koshtorys_needs[kekv] > 0:
            print(f"   КЕКВ {kekv}: {koshtorys_needs[kekv]:.2f} грн")
    print(f"   {'─' * 30}")
    print(f"   РАЗОМ:      {total_need:.2f} грн")

    # БАЛАНС ПО КОЖНОМУ КЕКВ
    print(f"\n⚖️  БАЛАНС ПО КОЖНОМУ КЕКВ:")
    total_deficit = 0
    total_surplus = 0

    for kekv in ['2210', '2240', '2250']:
        ostacha = total_ostacha.get(kekv, 0)
        potreba = koshtorys_needs.get(kekv, 0)
        diff = ostacha - potreba

        if ostacha > 0 or potreba > 0:
            if diff >= 0:
                print(f"   КЕКВ {kekv}: {ostacha:.2f} - {potreba:.2f} = +{diff:.2f} (надлишок)")
                total_surplus += diff
            else:
                print(f"   КЕКВ {kekv}: {ostacha:.2f} - {potreba:.2f} = {diff:.2f} (не вистачає)")
                total_deficit += abs(diff)

    print(f"   {'─' * 50}")

    overall_diff = total_ostacha_suma - total_need
    if overall_diff >= 0:
        print(f"   ✅ ЗАГАЛЬНИЙ БАЛАНС: +{overall_diff:.2f} грн (надлишок)")
    else:
        print(f"   ⚠️  ЗАГАЛЬНИЙ БАЛАНС: {overall_diff:.2f} грн (не вистачає)")

    if total_deficit > 0 and total_surplus > 0:
        print(f"\n   💡 Примітка: Є надлишки по одних КЕКВ та дефіцит по інших")

    # КРОК 5: Розрахунок РЕАЛЬНОЇ остачі
    print("\n" + "=" * 80)
    print("КРОК 5: Розрахунок реальної остачі")
    print("=" * 80)

    # РЕАЛЬНА ОСТАЧА = Забрали з Розрахунку - Потреба з Кошторису
    real_ostacha = {}
    print(f"\n💎 РЕАЛЬНА ОСТАЧА (для синіх рядків):")

    for kekv in ['2210', '2240', '2250']:
        zabrano = total_ostacha.get(kekv, 0)  # Скільки забрали
        potreba = koshtorys_needs.get(kekv, 0)  # Скільки треба
        ostacha = zabrano - potreba  # Реальна остача

        if ostacha > 0:
            real_ostacha[kekv] = ostacha
            print(f"   КЕКВ {kekv}: {zabrano:.2f} - {potreba:.2f} = {ostacha:.2f} грн")
        else:
            real_ostacha[kekv] = 0

    total_real_ostacha = sum(real_ostacha.values())
    if total_real_ostacha > 0:
        print(f"   {'─' * 50}")
        print(f"   РАЗОМ ОСТАЧА: {total_real_ostacha:.2f} грн")
    else:
        print(f"   (Немає остачі - все використано)")

    # КРОК 6: Запис в Зміни
    print("\n" + "=" * 80)
    print("КРОК 6: Запис даних у файл Зміни.xlsx")
    print("=" * 80)

    insert_data_to_zminy(
        zminy_file,
        koshtorys_data,
        zahody_rozrahunok,
        real_ostacha,  # ← ПЕРЕДАЄМО РЕАЛЬНУ ОСТАЧУ!
        rozrahunok_file  # ← ПЕРЕДАЄМО ШЛЯХ ДО РОЗРАХУНКУ!
    )

    print("\n" + "=" * 80)
    print("🎉 ПРОГРАМА ЗАВЕРШЕНА УСПІШНО!")
    print("=" * 80)
    print(f"\n📁 Перевірте файл '{zminy_file}'")
    print(f"\n💡 Рядки з остачею виділено блакитним кольором")


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print(f"\n\n⚠️  Програму перервано користувачем")
    except Exception as e:
        print(f"\n❌ КРИТИЧНА ПОМИЛКА: {e}")
        import traceback

        traceback.print_exc()
    finally:
        input("\nНатисніть Enter для закриття...")