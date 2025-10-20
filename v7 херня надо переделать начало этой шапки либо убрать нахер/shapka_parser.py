import openpyxl
from openpyxl.utils import get_column_letter
import json


def parse_excel_with_formatting(filename):
    """
    Парсит Excel файл с сохранением только значимого форматирования
    """
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    # Структура для хранения всей информации
    data = {
        'sheet_name': ws.title,
        'dimensions': {
            'max_row': ws.max_row,
            'max_column': ws.max_column
        },
        'merged_cells': [],
        'column_widths': {},
        'row_heights': {},
        'cells': []
    }

    # Запоминаем объединенные ячейки
    for merged_range in ws.merged_cells.ranges:
        data['merged_cells'].append(str(merged_range))
        print(f"Объединенные ячейки: {merged_range}")

    # Запоминаем ширину колонок (только если отличается от стандартной)
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        width = ws.column_dimensions[col_letter].width
        if width and width != 8.43:  # 8.43 - стандартная ширина
            data['column_widths'][col_letter] = width

    # Запоминаем высоту строк (только если отличается от стандартной)
    for row in range(1, ws.max_row + 1):
        height = ws.row_dimensions[row].height
        if height and height != 15:  # 15 - стандартная высота
            data['row_heights'][row] = height

    # Парсим каждую ячейку с полным форматированием
    print("\n=== СОДЕРЖИМОЕ И ФОРМАТИРОВАНИЕ ЯЧЕЕК ===\n")

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            col_letter = get_column_letter(col)
            cell_address = f"{col_letter}{row}"

            # Проверяем, является ли ячейка частью объединенной области
            is_merged = False
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    is_merged = True
                    break

            # Базовая информация о ячейке
            cell_info = {
                'address': cell_address,
                'row': row,
                'column': col
            }

            # Добавляем значение только если оно есть
            if cell.value is not None:
                cell_info['value'] = cell.value

            # Добавляем флаг объединения только если ячейка объединена
            if is_merged:
                cell_info['is_merged'] = True

            # Шрифт - сохраняем только отличия от стандартного
            if cell.font:
                font_info = {}
                if cell.font.name and cell.font.name != 'Calibri':
                    font_info['name'] = cell.font.name
                if cell.font.size and cell.font.size != 11:
                    font_info['size'] = cell.font.size
                if cell.font.bold:
                    font_info['bold'] = True
                if cell.font.italic:
                    font_info['italic'] = True
                if cell.font.underline and cell.font.underline != 'none':
                    font_info['underline'] = cell.font.underline

                if cell.font.color:
                    if hasattr(cell.font.color, 'rgb') and cell.font.color.rgb != '00000000':
                        font_info['color'] = cell.font.color.rgb
                    elif hasattr(cell.font.color, 'theme'):
                        font_info['color'] = f"theme:{cell.font.color.theme}"

                if font_info:
                    cell_info['font'] = font_info

            # Выравнивание - сохраняем только если есть
            if cell.alignment:
                alignment_info = {}
                if cell.alignment.horizontal and cell.alignment.horizontal != 'general':
                    alignment_info['horizontal'] = cell.alignment.horizontal
                if cell.alignment.vertical and cell.alignment.vertical != 'bottom':
                    alignment_info['vertical'] = cell.alignment.vertical
                if cell.alignment.wrap_text:
                    alignment_info['wrap_text'] = True
                if cell.alignment.text_rotation:
                    alignment_info['text_rotation'] = cell.alignment.text_rotation
                if cell.alignment.indent:
                    alignment_info['indent'] = cell.alignment.indent

                if alignment_info:
                    cell_info['alignment'] = alignment_info

            # Границы - сохраняем только если есть
            if cell.border:
                border_info = {}
                if cell.border.left and cell.border.left.style:
                    border_info['left'] = cell.border.left.style
                if cell.border.right and cell.border.right.style:
                    border_info['right'] = cell.border.right.style
                if cell.border.top and cell.border.top.style:
                    border_info['top'] = cell.border.top.style
                if cell.border.bottom and cell.border.bottom.style:
                    border_info['bottom'] = cell.border.bottom.style

                if border_info:
                    cell_info['border'] = border_info

            # Заливка - сохраняем только если есть цвет
            if cell.fill and cell.fill.patternType and cell.fill.patternType != 'none':
                fill_info = {'pattern_type': cell.fill.patternType}

                if cell.fill.fgColor:
                    if hasattr(cell.fill.fgColor, 'rgb') and cell.fill.fgColor.rgb != '00000000':
                        fill_info['fg_color'] = cell.fill.fgColor.rgb
                    elif hasattr(cell.fill.fgColor, 'theme'):
                        fill_info['fg_color'] = f"theme:{cell.fill.fgColor.theme}"

                if cell.fill.bgColor:
                    if hasattr(cell.fill.bgColor, 'rgb') and cell.fill.bgColor.rgb != '00000000':
                        fill_info['bg_color'] = cell.fill.bgColor.rgb
                    elif hasattr(cell.fill.bgColor, 'theme'):
                        fill_info['bg_color'] = f"theme:{cell.fill.bgColor.theme}"

                if len(fill_info) > 1:  # Если есть что-то кроме pattern_type
                    cell_info['fill'] = fill_info

            # Сохраняем только ячейки с содержимым или форматированием
            if len(cell_info) > 3:  # Больше чем просто address, row, column
                # Выводим информацию
                print(f"\n📍 Ячейка {cell_address}:")
                if 'value' in cell_info:
                    print(f"   Значение: {cell_info['value']}")
                if cell_info.get('is_merged'):
                    print(f"   ⚠️ Часть объединенной области")
                if cell_info.get('font', {}).get('bold'):
                    print(
                        f"   Жирный шрифт: {cell_info['font'].get('name', 'Calibri')}, размер {cell_info['font'].get('size', 11)}")
                if cell_info.get('alignment', {}).get('horizontal'):
                    print(f"   Выравнивание: {cell_info['alignment']['horizontal']}")
                if cell_info.get('fill', {}).get('fg_color'):
                    print(f"   Цвет заливки: {cell_info['fill']['fg_color']}")

                data['cells'].append(cell_info)

    # Сохраняем всё в JSON для дальнейшего использования
    with open('excel_format_data.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)

    print("\n" + "=" * 50)
    print(f"✅ Файл полностью проанализирован!")
    print(f"📊 Всего строк: {data['dimensions']['max_row']}")
    print(f"📊 Всего колонок: {data['dimensions']['max_column']}")
    print(f"🔗 Объединенных областей: {len(data['merged_cells'])}")
    print(f"💾 Сохранено ячеек с данными: {len(data['cells'])}")
    print(f"💾 Данные сохранены в excel_format_data.json")
    print("=" * 50)

    return data




if __name__ == '__main__':
    try:
        result = parse_excel_with_formatting('шапка.xlsx')
    except FileNotFoundError:
        print("❌ Файл 'шапка.xlsx' не найден!")
    except Exception as e:
        print(f"❌ Ошибка: {e}")

    finally:
        input("\nНатисніть Enter для закриття...")